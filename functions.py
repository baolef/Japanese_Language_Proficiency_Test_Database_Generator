# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Project : Japanese Language Proficiency Test Database Generator
# @Time    : 2020/2/7 19:10
# @Author  : Fang Baole
# @Email   : fbl718@sjtu.edu.cn
import traceback

import openpyxl


def replace(ws, olds, new):
    """Replaces repeated old with new.

    The function replaces element in olds with new in ws.

    Args:
        ws: A work sheet for output.
        olds: A list of characters needed to be replaced.
        new: A string to replace.

    Returns:
        None
    """
    rows = ws.max_row
    cols = ws.max_column
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            text = ws.cell(row=row, column=col).value
            if text != None and isinstance(text, str):
                for old in olds:
                    times = text.count(old)
                    for length in range(times, 0, -1):
                        old_string = repeated_string(old, length)
                        text = text.replace(old_string, new)
                ws.cell(row=row, column=col).value = text


def repeated_string(char, times):
    """Returns a string of repeated times of char.

    The function returns a string of times times of char.

    Args:
        char: A string that needs to be repeated.
        times: An integer showing the times needed to repeat.

    Returns:
        A string that has been repeated by times times.
    """
    string = ''
    for i in range(times):
        string = string + char
    return string


def read(file, answer_list, ws, wb, y, m, l, mark, output_name):
    """Reads the test paper and writes in the work sheet.

    The function reads the test paper and writes all the titles, questions and options in the work sheet.
    The test paper is read paragraph by paragraph. When a paragraph is read, it is decided whether it belongs to:
    1. Listening. If the key word listening is read, all paragraphs after belongs to listening, else not.
    2. Option. If it is not listening, then we first decide whether is is an option, which is decided by whether there is a question before.
    3. Title. If it is not an option, then we decide whether it is a title, which is decided by whether its text begins with certain key words.
    4. Question. If it is not a title, then we decide whether it is a question, which is decided by check_number function.
    5. Nothing. If it is none of above, then skip.
    Once the type of paragraph is decided, corresponding information is written into the work sheet.
    Be aware that:
    1. The listening part is written by the write_listening function.
    2. The answers are written and collected before this function because certain features need its structure.

    Args:
        file: A Word document of the answer file.
        answer_list: A list of answers.
        ws: A work sheet for output.
        wb: A work book for output.
        y: An integer that describes the year of the test.
        m: An integer that describes the month of the test.
        l: An integer that describes the level of the test.
        mark: An integer that describes the type of the test.
        output_name: A string that describes the name of the work sheet.

    Returns:
        None
    """
    line = 2
    options = 0
    element_list = ['１', '２', '３', '４']
    # element_list=['1','2','3','4']
    string = ''
    flag = False
    judge = False
    listening = False
    options_list = []
    for para in file.paragraphs:
        try:
            if listening:
                try:
                    if para.runs[0].bold:
                        judge = not para.text.startswith('例')
                    elif judge:
                        temp = underline_check(para)
                        if temp.split().__len__() % 2 == 1:
                            for element in element_list:
                                temp = temp.replace(element, ' ' + element + ' ')
                        str_list = temp.split()
                        # if str_list.__len__()==8:
                        #     options_list.append()
                        for element in str_list:
                            if element in element_list:
                                if flag:
                                    options_list.append(string[:-1])
                                    flag = False
                                string = str(element_list.index(element) + 1) + '.'
                                flag = True
                            else:
                                string = string + element + ' '
                        if flag:
                            options_list.append(string[:-1])
                            flag = False
                except:
                    continue
            else:
                listening = "听力" in para.text or "答题卡" in para.text
                if options > 0:
                    temp = underline_check(para)
                    if temp.split().__len__() % 2 == 1:
                        for element in element_list:
                            temp = temp.replace(element, ' ' + element + ' ')
                    str_list = temp.split()
                    if str_list[0] not in element_list:
                        if options == 4:
                            ws['B' + str(line)] = ws['B' + str(line)].value + '\n' + underline_check(para)
                            continue
                        else:
                            wb.save(output_name)
                            print('Error happens at line ' + str(line))
                            quit(-1)
                    for element in str_list:
                        if element in element_list:
                            if flag:
                                ws['C' + str(line)] = string[:-1]
                                line = line + 1
                                options = options - 1
                                flag = False
                            string = str(element_list.index(element) + 1) + '.'
                            flag = True
                        else:
                            string = string + element + ' '
                    if flag:
                        ws['C' + str(line)] = string[:-1]
                        line = line + 1
                        options = options - 1
                        flag = False
                elif para.text.startswith('問題'):
                    ws['A' + str(line)] = underline_check(para)
                elif check_number(para.text):
                    options = 4
                    write_information(line, ws, y, m, l, mark)
                    if check_potential_option(para.text):
                        ws['B' + str(line)] = para.text[0:para.text.find('.') + 1]
                        temp = underline_check(para)
                        for element in element_list:
                            temp = temp.replace(element, ' ' + element + ' ')
                        str_list = temp.split()
                        for element in str_list:
                            if element in element_list:
                                if flag:
                                    ws['C' + str(line)] = string[:-1]
                                    line = line + 1
                                    options = options - 1
                                    flag = False
                                string = str(element_list.index(element) + 1) + '.'
                                flag = True
                            else:
                                string = string + element + ' '
                        if flag:
                            ws['C' + str(line)] = string[:-1]
                            line = line + 1
                            options = options - 1
                            flag = False
                    else:
                        ws['B' + str(line)] = underline_check(para)
        except Exception as e:
            print('Program stops at line ' + str(para.text))
            print(traceback.format_exc())
    write_listening(line, options_list, answer_list, ws, y, m, l, mark)


def write_listening(line, options_list, answer_list, ws, y, m, l, mark):
    """Writes the listening part of the test into the work sheet.

    The function writes the listening part of the test into the work sheet.
    It is done by first getting the answer list to know the structure of the listening part.
    With the structure, using the known format of the test paper to write the titles and questions,
    during which the options are also written from the option list.

    Args:
        line: An integer that describes the current line to write in.
        options_list: A list of options.
        answer_list: A list of answers.
        ws: A work sheet for output.
        y: An integer that describes the year of the test.
        m: An integer that describes the month of the test.
        l: An integer that describes the level of the test.
        mark: An integer that describes the type of the test.

    Returns:
        None
    """
    title_list = ['問題１ 問題１では、まず質問を聞いてください。それから話を聞いて、問題用紙の１から４の中から、最もよいものを一つ選んでください。',
                  '問題2 問題2では、まず質問を聞いてください。そのあと、問題用紙のせんたくしを読んでください。読む時間があります。それから話を聞いて、問題用紙の１から４の中から、最もよいものを一つ選んでください。',
                  '問題3 問題3では、問題用紙に何も印刷されていません。この問題は、全体としてどんな内容かを聞く問題です。話の前に質問はありません。まず話を聞いてください。それから、質問とせんたくしを聞いて、１から４の中から、最もよいものを一つ選んでください。 ',
                  '問題4 問題4では、問題用紙に何も印刷されていません。まず文を聞いてください。それから、それに対する返事を聞いて、１から３の中から、最もよいものを一つ選んでください。 ',
                  '問題5 問題5では、長めの話を聞きます。この問題には練習はありません。メモをとってもかまいません。\n1番、2番 問題用紙に何も印刷されていません。まず話を聞いてください。それから、質問とせんたくしを聞いて、１から４の中から、最もよいものを一つ選んでください。 '
                  '問題5 問題5では、長めの話を聞きます。この問題には練習はありません。メモをとってもかまいません。\n３番 まず話を聞いてください。それから、二つの質問を聞いて、それぞれ問題用紙の１から４の中から、最もよいものを一つ選んでください。']
    for temp_list in answer_list[0:2]:
        ws['A' + str(line)] = title_list.pop(0)
        num = 0
        for temp in temp_list:
            num = num + 1
            ws['B' + str(line)] = str(num) + '番'
            write_information(line, ws, y, m, l, mark)
            for i in range(0, 4):
                ws['C' + str(line)] = options_list.pop(0)
                line = line + 1
    for temp_list in answer_list[2:]:
        ws['A' + str(line)] = title_list.pop(0)
        num = 0
        for temp in temp_list:
            num = num + 1
            ws['B' + str(line)] = str(num) + '番'
            write_information(line, ws, y, m, l, mark)
            for i in range(1, 5):
                ws['C' + str(line)] = str(i) + '.'
                line = line + 1
    while options_list.__len__() > 0:
        line = line - 1
        ws['C' + str(line)] = options_list.pop()


def write_information(line, ws, y, m, l, mark):
    """Write other information, like test type, year, month and question type into the work sheet.

    The function writes test type, year, month and question type into the work sheet when a question is written.
    
    Args:
        line: An integer that describes the current line to write in.
        ws: A work sheet for output.
        y: An integer that describes the year of the test.
        m: An integer that describes the month of the test.
        l: An integer that describes the level of the test.
        mark: An integer that describes the type of the test.

    Returns:
        None
    """
    # 类型
    ws['E' + str(line)] = int(mark)
    # 年份
    ws['F' + str(line)] = int(y)
    # 月份
    ws['G' + str(line)] = int(m)
    # 题型
    ws['J' + str(line)] = 1


def check_number(text):
    """Checks whether the text is a question.

    The function checks whether the input text is a question by examining whether the text begins with "num.".

    Args:
        text: A string that describes a potential question.

    Returns:
        A bool showing whether the input text is a question.
    """
    index = text.find('.')
    if index < 0:
        index = text.find('．')
    try:
        num = int(text[0:index])
    except:
        return False
    else:
        return True


def check_potential_option(text):
    """Check whether the text is a potential option.

    When processing the test paper file, there is likely to have questions and options in the same paragraph.
    The function checks whether there is an option in the question text.

    Args:
        text: A string that describes the potential question.

    Returns:
        A bool showing whether there is an option in the question text.
    """
    text = text.replace(' ', '')
    index = text.find('.')
    return text[index + 1] == '１'


def underline_check(para):
    """Convert the underline part of the file to specific format.

    The function returns a processed text from paragraph, whose underline style is converted.

    Args:
        para: A paragraph
    """
    runs = para.runs
    str = ''
    star = False
    for run in runs:
        if run.underline:
            if not run.text.replace(' ', ''):
                if star:
                    star = False
                else:
                    str = str + '<under2></under2>'
            elif run.text == '★':
                str = str[:-17] + '<under>' + run.text + '</under>'
                star = True
            else:
                str = str + '<under>' + run.text + '</under>'
        else:
            str = str + run.text
    str = str.replace('> ', '>')
    str = str.replace(' <', '<')
    return str


def get_sheetname(filename):
    """Gets the sheet name of the output sheet.

    The function gets the sheet name of the ouput sheet from the filename of the test paper.
    It also extracts the year, month and level of the test from the filename.

    Args:
        filename: A string that describes the name of the test paper file.

    Returns:
        A string of the sheet name and three integers of the year, month and level of the test.
    """
    try:
        year = filename.find('年')
        month = filename.find('月')
        level = filename.find('N')
        y = filename[year - 4:year]
        m = filename[year + 1:month]
        l = filename[level + 1]
        return y + '-' + m + '-N' + l, y, m, l
    except:
        print("Can't find the sheet name. Please input the sheet name manually.")
        y = input('Year: ')
        m = input('Month: ')
        l = input('Level: ')
        sheetname = y + '-' + m + '-N' + l
        return sheetname, y, m, l


def add_sheet(filename, output_name):
    """Creates a work sheet for output.

    The function creates a work sheet output.

    Args:
        filename: A string that describes the name of the test paper file.
        output_name: A string that describes the name of the work sheet.

    Returns:
        A work sheet and a work book of the output file, the year, month and level of the test.
    """
    sheetname, y, m, l = get_sheetname(filename)
    try:
        wb = openpyxl.load_workbook(output_name)
    except:
        print('Creating workbook "' + output_name + '"...')
        wb = openpyxl.Workbook()
        print('Creation succeed.')
    while sheetname in wb.sheetnames:
        sheetname = input(
            'Sheet name "' + sheetname + '" is already in the workbook "' + output_name + '". Please enter a new sheet name: ')
    ws = wb.active
    if ws.title.title() == 'Sheet':
        ws.title = sheetname
    else:
        ws = wb.create_sheet(sheetname)
    print('Creating work sheet "' + sheetname + '"...')
    headers = ['题干', '问题', '选项', '是否正确答案', '类型', '年份', '月份', '分类1', '分类2', '题型']
    ws.append(headers)
    print('Creation succeed.')
    return ws, wb, y, m, l


def read_answer(answer, ws):
    """Reads the answer file and writes the answers in the work sheet.

    The function processes the answer file paragraph by paragraph to find the one that contains answers.
    Once the paragraph is found, it is processed by get_answer_list function to get the answer list.
    The answer list is then written in the work sheet by openpyxl.

    Args:
        answer: A Word document of the answer file.
        ws: A work sheet for output.

    Returns:
        A list of all answers.
    """
    answer_list = []
    line = 0
    question_type = 0
    type_list = []
    for para in answer.paragraphs:
        if para.text.startswith('問題'):
            temp = get_answer_list(para.text)
            answer_list.append(temp)
            for i in range(len(temp)):
                type_list.append(question_type)
        elif para.text.startswith('文字'):
            question_type = 1
        elif para.text.startswith('文法'):
            question_type = 2
        elif para.text.startswith('読解'):
            question_type = 3
        elif para.text.startswith('聴解'):
            question_type = 4
    for group in answer_list:
        for ans in group:
            ws['D' + str(line * 4 + ans + 1)] = 1
            ws['H' + str(line * 4 + 2)] = type_list.pop(0)
            line = line + 1
    return answer_list


def get_answer_list(text):
    """Gets a list of answer from the input text.

    The function splits the text by the division mark to get the answer list.

    Args:
        text: A string of the line got from the answer file.

    Returns:
        A list of answers extracted from the text.
    """
    # mark=get_most_char(text.replace(' ',''))
    while not text[-1].isnumeric():
        text = text[:-1]
    mark = text[-2]
    text = text.replace(mark, ' ')
    text = text.replace('：', ' ')
    str_list = text.split()
    answer_list = []
    # if mark in text:
    #     str_list = text.replace(' ', '').split(mark)
    while str_list.__len__() > 0:
        element = str_list.pop(0)
        if len(element) == 1:
            try:
                temp = int(element)
            except:
                continue
            else:
                if temp in range(1, 5):
                    answer_list.append(temp)
    return answer_list


def get_most_char(text):
    """Gets the most frequent non-integer character in a string.

    The function returns the most frequent character in a string if it is not an integer by sorting dictionary.
    It is designed to find the division mark between each answer which is different due to different answer formats,
    like "." and so on, so that get_answer_list function can read the answers
    However, the function is not used in this program because there sometimes occurs that there are only two answers
    in one line, resulting in only one division mark, which may not be recognized by this function since it may not be
    the character that appears most frequent.

    Args:
        text: A string from which the function finds the most frequent non-integer character.

    Returns:
        A non-integer character that appears most frequent in the input string.

    """
    dic = {}
    for char in text:
        if char in dic:
            dic[char] += 1
        else:
            dic[char] = 1
    d_order = sorted(dic.items(), key=lambda x: x[1], reverse=True)
    for item in d_order:
        try:
            int(item[0])
        except:
            print(item[0])
            return item[0]
        else:
            continue


def check_type(filename):
    """Decides the type of the test paper.

    The function decides the type of the test paper file by its filename and returns its type.

    Args:
        filename: A string that describes the name of the test paper file.

    Returns:
        An integer that represents the type of the input test paper file.
    """
    if "真题" in filename:
        return 1
    elif "模考" in filename:
        return 2
    else:
        return ''
